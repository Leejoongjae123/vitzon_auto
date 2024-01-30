import openpyxl
import pandas as pd
import requests
import xmltodict
import pprint
from bs4 import BeautifulSoup
import json
import os
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import datetime
import time
import re
import boto3
import xml.etree.ElementTree as ET
import codecs

# 정렬 키를 만들기 위한 함수
def sorting_key(item):
    return (item["레벨1"], item["레벨2"], item["레벨3"], item["레벨4"])


def GetLogin():
    # ---------------cURL 및 CURL CONVERTER이용하여 작성하는 부분
    cookies = {
        'mroId': 'KAPRL2023',
        '_fwb': '100IedV0seCBo5GfsIUMfUf.1706023209455',
        'grb_ck@39723607': '6f612887-4f7a-c7f5-d68f-bec47f8705ca',
        '_ga': 'GA1.1.548138226.1706023210',
        'grb_ui@39723607': '92db61e2-4953-bf22-0bda-4f8451129bc4',
        'grb_recent_goods@39723607': '55112%2C55112%2C236333%2C236333%2C236333',
        'grb_id_permission@39723607': 'success',
        'grb_ip_permission@39723607': 'success',
        'grb_dynamic_list@': 'checked%2C%5B%5D',
        'JSESSIONID': '1CFD04C2FB339FE3AF7F38F428F56037',
        'wcs_bt': '1be955ead4801f:1706369896',
        '_ga_HSPV75KRK1': 'GS1.1.1706369846.8.1.1706369905.0.0.0',
    }

    headers = {
        'authority': 'vitsonmro.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'cache-control': 'max-age=0',
        'content-type': 'application/x-www-form-urlencoded',
        # 'cookie': 'mroId=KAPRL2023; _fwb=100IedV0seCBo5GfsIUMfUf.1706023209455; grb_ck@39723607=6f612887-4f7a-c7f5-d68f-bec47f8705ca; _ga=GA1.1.548138226.1706023210; grb_ui@39723607=92db61e2-4953-bf22-0bda-4f8451129bc4; grb_recent_goods@39723607=55112%2C55112%2C236333%2C236333%2C236333; grb_id_permission@39723607=success; grb_ip_permission@39723607=success; grb_dynamic_list@=checked%2C%5B%5D; JSESSIONID=1CFD04C2FB339FE3AF7F38F428F56037; wcs_bt=1be955ead4801f:1706369896; _ga_HSPV75KRK1=GS1.1.1706369846.8.1.1706369905.0.0.0',
        'origin': 'https://vitsonmro.com',
        'referer': 'https://vitsonmro.com/mro/login.do',
        'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    }

    data = {
        'loc': 'mro',
        'custId': 'KAPRL2023',
        'custPw': 'a123456',
    }

    session = requests.session()
    response = session.post('https://vitsonmro.com/mro/login', cookies=cookies, headers=headers, data=data)

    print("status_code:", response.status_code)
    print("status_code:", response.headers)

    cookies = {
        'mroId': 'KAPRL2023',
    }

    headers = {
        'authority': 'vitsonmro.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'cache-control': 'max-age=0',
        # 'cookie': 'mroId=KAPRL2023; _fwb=100IedV0seCBo5GfsIUMfUf.1706023209455; grb_ck@39723607=6f612887-4f7a-c7f5-d68f-bec47f8705ca; _ga=GA1.1.548138226.1706023210; grb_ui@39723607=92db61e2-4953-bf22-0bda-4f8451129bc4; grb_id_permission@39723607=success; grb_ip_permission@39723607=success; grb_dynamic_list@=checked%2C%5B%5D; grb_recent_goods@39723607=55112%2C236333%2C236333%2C236333%2C473637; wcs_bt=1be955ead4801f:1706370282; _ga_HSPV75KRK1=GS1.1.1706369846.8.1.1706370291.0.0.0; JSESSIONID=772364703B675C228219586B53B671CF',
        'referer': 'https://vitsonmro.com/mro/login.do',
        'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    }

    response = session.get('https://vitsonmro.com/mro/shop/productDetail.do?productCode=473637', cookies=cookies,
                           headers=headers)
    print("response.headers:", response.headers, "/ response.headers_TYPE:", type(response.headers))
    sessionid = response.headers['Set-Cookie']
    # '='와 ';' 사이의 값을 추출하는 정규 표현식
    sessionid = re.findall(r'=(.*?);', sessionid)[0]
    print("sessionid:", sessionid, "/ sessionid_TYPE:", type(sessionid))
    return sessionid
def GetCS():
    # XML 파일 위치
    xml_url = "https://myvitzonbucket.s3.us-east-2.amazonaws.com/change.xml"

    # 등록접근주소 (base URL)
    base_url = "https://sbadmin15.sabangnet.co.kr/RTL_API/xml_goods_info.html"

    # 전체 URL 구성
    full_url = f"{base_url}?xml_url={xml_url}"

    print("full_url:",full_url,"/ full_url_TYPE:",type(full_url))
    # GET 요청 보내기
    response = requests.post(full_url)

    # 응답 상태 코드 및 내용 출력
    print("Status Code:", response.status_code)
    print("Response Body:", response.text)
    results=xmltodict.parse(response.text)
    pprint.pprint(results)
def AddCategory():
    # XML 파일 위치
    xml_url = "https://myvitzonbucket.s3.us-east-2.amazonaws.com/category.xml"

    # 등록접근주소 (base URL)
    base_url = "https://sbadmin15.sabangnet.co.kr/RTL_API/xml_category_info2.html"

    # 전체 URL 구성
    full_url = f"{base_url}?xml_url={xml_url}"

    print("full_url:",full_url,"/ full_url_TYPE:",type(full_url))
    # GET 요청 보내기
    response = requests.post(full_url)

    # 응답 상태 코드 및 내용 출력
    print("Status Code:", response.status_code)
    print("Response Body:", response.text)
def ChangeProduct():
    # XML 파일 위치
    xml_url = "https://myvitzonbucket.s3.us-east-2.amazonaws.com/change.xml"

    # 등록접근주소 (base URL)
    base_url = "https://sbadmin15.sabangnet.co.kr/RTL_API/xml_goods_info.html"

    # 전체 URL 구성
    full_url = f"{base_url}?xml_url={xml_url}"

    print("full_url:",full_url,"/ full_url_TYPE:",type(full_url))
    # GET 요청 보내기
    response = requests.post(full_url)

    # 응답 상태 코드 및 내용 출력
    print("Status Code:", response.status_code)
    # print("Response Body:", response.text)
    if response.text.find("SUCCESS")>=0:
        print("★반영완료★")
    else:
        print("response.text:",response.text,"/ response.text_TYPE:",type(response.text))
def makeCategory():
    data_dict_list = []
    for i in range(1,5):
        df=pd.read_csv('list{}.csv'.format(i))
        # V, W, X, Y 열의 정보를 딕셔너리로 저장
        # 21번째 열부터 24번째 열까지의 정보를 딕셔너리로 만들기

        # 22번째 열(인덱스 21)의 값이 비어 있고, 25번째 열(인덱스 24)의 값이 있는 행 필터링
        filtered_rows = df[df.iloc[:, 22].isna() & df.iloc[:, 23].notna()]

        # 결과 출력
        print(filtered_rows)

        for index, row in df.iterrows():
            info_dict = {}
            for col_num in range(21, 25):  # 21번째 열부터 24번째 열까지
                column_name = df.columns[col_num]  # 열 이름 가져오기
                value = row[col_num] if not pd.isna(row[col_num]) else ''  # NaN이면 빈 문자열 처리
                info_dict[column_name] = value
            data_dict_list.append(info_dict)

    print(len(data_dict_list))
    # data_dict_list에는 21번째 열부터 24번째 열까지의 정보가 순차적으로 딕셔너리로 저장됩니다.
    with open('data_dict_list.json', 'w',encoding='utf-8-sig') as f:
        json.dump(data_dict_list, f, indent=2,ensure_ascii=False)



    # 중복 제거를 위해 set을 사용
    unique_data_set = set()

    # 중복을 제거하고 유일한 요소만 unique_data_set에 추가
    for item in data_dict_list:
        item_values = tuple(item.values())
        unique_data_set.add(item_values)

    # 중복이 제거된 리스트로 변환
    unique_data_list = [dict(zip(item.keys(), item_values)) for item_values in unique_data_set]

    # print(unique_data_list)
    print(len(unique_data_list))
    pprint.pprint(unique_data_list)
def makeCategory2():
    with open ('unique_data_list.json', "r",encoding='utf-8-sig') as f:
        unique_data_list = json.load(f)

    # 정렬 수행
    unique_data_list = sorted(unique_data_list, key=sorting_key)

    with open('unique_data_list.json', 'w',encoding='utf-8-sig') as f:
        json.dump(unique_data_list, f, indent=2,ensure_ascii=False)

    # 레벨별 코드 부여
    level1_dict = {}
    level2_dict = {}
    level3_dict = {}
    level4_dict = {}

    for item in unique_data_list:
        level1_value = item["레벨1"]
        level2_value = item["레벨2"]
        level3_value = item["레벨3"]
        level4_value = item["레벨4"]

        if level1_value:
            if level1_value in level1_dict:
                item["레벨1_코드"] = level1_dict[level1_value]
            else:
                level1_dict[level1_value] = f"A{len(level1_dict) + 1}"
                item["레벨1_코드"] = level1_dict[level1_value]

        if level2_value:
            if level2_value in level2_dict:
                item["레벨2_코드"] = level2_dict[level2_value]
            else:
                level2_dict[level2_value] = f"B{len(level2_dict) + 1}"
                item["레벨2_코드"] = level2_dict[level2_value]

        if level3_value:
            if level3_value in level3_dict:
                item["레벨3_코드"] = level3_dict[level3_value]
            else:
                level3_dict[level3_value] = f"C{len(level3_dict) + 1}"
                item["레벨3_코드"] = level3_dict[level3_value]

        if level4_value:
            if level4_value in level4_dict:
                item["레벨4_코드"] = level4_dict[level4_value]
            else:
                level4_dict[level4_value] = f"D{len(level4_dict) + 1}"
                item["레벨4_코드"] = level4_dict[level4_value]

    print(unique_data_list)

    wb=openpyxl.Workbook()
    ws=wb.active
    for unique_data in unique_data_list:
        data=[""]*12
        try:
            data[0]=unique_data['레벨1_코드']
        except:
            data[0]=""

        try:
            data[1]=unique_data['레벨1']
        except:
            data[1]=""

        try:
            data[3]=unique_data['레벨2_코드']
        except:
            data[3]=""
        try:
            data[4]=unique_data['레벨2']
        except:
            data[4]=""

        try:
            data[6]=unique_data['레벨3_코드']
        except:
            data[6]=""
        try:
            data[7]=unique_data['레벨3']
        except:
            data[7]=""

        try:
            data[9]=unique_data['레벨4_코드']
        except:
            data[9]=""
        try:
            data[10]=unique_data['레벨4']
        except:
            data[10]=""

        # data=[unique_data['레벨1_코드'],unique_data['레벨1'],"",unique_data['레벨2_코드'],unique_data['레벨2'],"",unique_data['레벨3_코드'],unique_data['레벨3'],"",unique_data['레벨4_코드'],unique_data['레벨4']]
        ws.append(data)
    wb.save('result.xlsx')

def MergeExcel():
    folder_path = 'test'
    files = os.listdir(folder_path)

    # 모든 데이터를 저장할 빈 DataFrame 생성
    all_data = pd.DataFrame()

    for index,file in enumerate(files):
        print("{}/{}번째 가져오기..".format(index+1,len(files)))
        file_path = os.path.join(folder_path, file)
        # 파일이 엑셀 파일인지 확인 (확장자가 .xlsx 또는 .xls)
        if os.path.isfile(file_path) and file_path.endswith(('.xlsx', '.xls')):
            # 4행부터 데이터를 읽음 (header=3)
            data = pd.read_excel(file_path, header=2,usecols=['상품명','상품약어','자체상품코드','물류처','상품상태','원가','판매가','대표이미지','상품상세설명','브랜드명'])
            # '상품상세설명' 열에서 <span>을 <p>로, </span>을 </p>로 변경
            data['상품상세설명'] = data['상품상세설명'].str.replace('<span', '<p', regex=False)
            data['상품상세설명'] = data['상품상세설명'].str.replace('</span>', '</p>', regex=False)
            # 읽은 데이터를 all_data에 추가
            all_data = pd.concat([all_data, data], ignore_index=True)

            # pprint.pprint(data)

    all_data.to_excel("totalResult.xlsx")
    wb=openpyxl.load_workbook('totalResult.xlsx')
    ws=wb.active
    ws["A1"]="순서"
    # ws["A1"] = "순서"
    # ws["D1"] = "자체상품코드"
    wb.save("totalResult.xlsx")
def GetGoogleSpreadSheet(dataList):
    scope = 'https://spreadsheets.google.com/feeds'
    json = 'credential_client.json'
    credentials = ServiceAccountCredentials.from_json_keyfile_name(json, scope)
    gc = gspread.authorize(credentials)
    sheet_url = 'https://docs.google.com/spreadsheets/d/18nYQTq9rotCYP2biQwY8KGzNHNL_XKJPdkMPrikC5EI/edit#gid=0'
    doc = gc.open_by_url(sheet_url)
    worksheet = doc.worksheet('시트1')
    #=================특정행의 정보 가져오기
    # cell_data = worksheet.acell('A1').value
    #=================전체정보가져오기
    # all_data=worksheet.get_all_records()
    #==================맨 밑행에 데이타 넣기
    print("송신시작")

    # 스프레드시트에 데이터 업로드

    # worksheet.update('A2', dataList)  # A1 셀부터 시작하여 데이터 업로드

    # 초기 시작 행을 지정 (예: 2행부터 시작)
    start_row = 2

    for data_chunk in dataList:
        # 데이터 업데이트할 셀 범위 결정
        end_row = start_row + len(data_chunk) - 1
        cell_range = f"A{start_row}:I{end_row}"  # 예시로 'A'열부터 'D'열까지 데이터를 넣는다고 가정

        # 데이터를 스프레드시트에 업데이트
        worksheet.update(cell_range, data_chunk)

        # 다음 데이터 청크를 위한 시작 행 업데이트
        start_row = end_row + 1
    print("구글스프레드작성완료")
def GetExcel():
    # CSV 파일을 읽습니다. 여기서 'your_file.csv'를 실제 파일 경로로 바꿔주세요.
    # 헤더네임 ['상품명','상품약어','자체상품코드\n[수정불가]','물류처ID','상품상태','원가','판매가','대표이미지']
    # df = pd.read_csv('totalResult.csv', usecols=[1, 2, 6, 11, 18, 24, 25, 31],header=None)
    df = pd.read_excel('totalResult.xlsx',index_col=False)

    # 두 번째 열부터 마지막 열까지 선택
    selected_data = df.iloc[:, 1:]
    # nan은 빈거로
    df = df.fillna("")


    df = df.to_dict(orient='record')
    # pprint.pprint(df)
    return df
def GetID(data,session):
    resultList = []
    pageCount = 1
    while True:
        cookies = {
            # 'grb_ck@39723607': 'eb03e8f0-fd89-1ca8-fec5-93f1b0a803fa',
            # '_ga': 'GA1.1.1775748282.1698133317',
            # 'grb_ui@39723607': '92db61e2-4953-bf22-0bda-4f8451129bc4',
            # 'grb_recent_goods@39723607': '381711%2C8501%2C382557%2C448294%2C85233',
            # 'grb_id_permission@39723607': 'success',
            # 'grb_ip_permission@39723607': 'success',
            # 'grb_dynamic_list@': 'checked%2C%5B%5D',
            'JSESSIONID': session,
            # 'wcs_bt': '1be955ead4801f:1701005537',
            # '_ga_HSPV75KRK1': 'GS1.1.1701005453.26.1.1701005537.0.0.0',
        }

        headers = {
            'authority': 'vitsonmro.com',
            'accept': 'application/json, text/javascript, */*; q=0.01',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'content-type': 'application/json',
            # 'cookie': 'grb_ck@39723607=eb03e8f0-fd89-1ca8-fec5-93f1b0a803fa; _ga=GA1.1.1775748282.1698133317; grb_ui@39723607=92db61e2-4953-bf22-0bda-4f8451129bc4; grb_recent_goods@39723607=381711%2C8501%2C382557%2C448294%2C85233; grb_id_permission@39723607=success; grb_ip_permission@39723607=success; grb_dynamic_list@=checked%2C%5B%5D; JSESSIONID=FAB29C22C2B8804D22C2492274A1917A; wcs_bt=1be955ead4801f:1701005537; _ga_HSPV75KRK1=GS1.1.1701005453.26.1.1701005537.0.0.0',
            'origin': 'https://vitsonmro.com',
            'referer': 'https://vitsonmro.com/mro/shop/productList.do?keyword=&productCode=&productNm=&standard=&brandNm=&subKeyword=&lvl1Cty=074060&lvl2Cty=&lvl3Cty=&dispCode=&statusCd1=&statusCd2=&categoryType=N&sorts=',
            'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            'x-requested-with': 'XMLHttpRequest',
        }

        json_data = {
            'keyword': '',
            'subKeyword': '',
            'productCode': '',
            'productNm': '',
            'brandNm': '',
            'standard': '',
            'categoryType': 'N',
            'lvl1Cty': data['categoryId'],
            'lvl2Cty': '',
            'lvl3Cty': '',
            'sorts': '',
            'engineYn': 'Y',
            'take': 60,
            'skip': 120,
            'page': pageCount,
            'pageSize': 60,
        }

        response = requests.post('https://vitsonmro.com/mro/shop/selectProductList.do', cookies=cookies,
                                 headers=headers,
                                 json=json_data)
        totalAmount = json.loads(response.text)['total']
        results = json.loads(response.text)['data']

        if len(results) == 0:
            print('페이지 없어서 종료')
            return resultList
        for result in results:
            # pprint.pprint(result)
            try:
                productNm=result['productNm']
            except:
                productNm=""
            # print("productNm:",productNm)
            try:
                productCode=result['productCode']
            except:
                productCode=""
            # print("productCode:",productCode)
            try:
                standard=result['standard']
            except:
                standard=""
            # print("standard:",standard)
            try:
                originPrice=int(round(int(result['rawUnitPrice'])*1.1,-1))
            except:
                originPrice=""
            # print("originPrice:",originPrice)
            try:
                salePrice=int(round(int(result['negoPrice'])*1.1,-1))
            except:
                salePrice=""
            # print("salePrice:",salePrice)
            
            try:
                brandNm=result['brandNm']
            except:
                brandNm=""
            print("brandNm:",brandNm)
            try:
                balance=result['onhand']
                if balance=="재고보유":
                    balance=2
                elif balance=="일시품절":
                    balance=3
            except:
                balance=""
            # print("balance:",balance)
            try:
                unit="VISTA"+result['eachAmount']
            except:
                unit=""
            # print("unit:",unit)
            try:
                pictureNm=result['pictureNm']
            except:
                pictureNm=""
            # print("pictureNm:",pictureNm)
            result = {'text': data['text'], 'productCode': productCode, 'categoryId': data['categoryId'],'pageCount': pageCount,
                      'productNm':productNm,'standard':standard,'originPrice':originPrice,'salePrice':salePrice,'balance':balance,'unit':unit,'pictureNm':pictureNm,'brandNm':brandNm}
            resultList.append(result)
            print("result:",result,"/ result_TYPE:",type(result))
            print("=============================================")
            if len(resultList) >= totalAmount:
                print('갯수충족')
                return resultList

        pageCount += 1
        print("data['text']:", data['text'], "/ data['text']_TYPE:", type(data['text']), len(data['text']))
        print("pageCOunt:", pageCount, "/ pageCOunt_TYPE:", type(pageCount), "현재갯수:", len(resultList))
        time.sleep(0.2)
        return resultList


def MakeXML(data,inputElem,statusType):
    timeNow=datetime.datetime.now().strftime("%Y%m%d")
    if statusType==True:
        data={"SABANG_GOODS_REGI":{
            'HEADER': {'SEND_COMPAYNY_ID': 'kaprl010',
                       'SEND_AUTH_KEY': 'AyTNrA19M8M653FX3MJ0CPBY0yG2J4STRbH',
                       'SEND_DATA': timeNow,
                       'SEND_GOODS_CD_RT': 'Y',
                       'RESULT_TYPE': 'XML'},
            'DATA': {'GOODS_NM': "<![CDATA[{}]]>".format(data['productNm']),
                     'GOODS_KEYWORD': "<![CDATA[{}]]>".format(data['standard']),
                     'COMPAYNY_GOODS_CD': "<![CDATA[vitson_{}]]>".format(str(data['productCode'])),
                     'STATUS': data['balance'],
                     'DPARTNER_ID': "<![CDATA[{}]]>".format(data['unit']),
                     'PARTNER_ID': "<![CDATA[vitsonmro]]>",
                     'IMG_PATH': "<![CDATA[{}]]>".format(data['pictureNm']),
                     'GOODS_COST': "<![CDATA[{}]]>".format(data['salePrice']),
                     'STOCK_USE_YN': 'N',
                     'GOODS_PRICE': "<![CDATA[{}]]>".format(data['originPrice']),
                     'GOODS_REMARKS': "<![CDATA[{}]]>".format(inputElem['상품상세설명']),
                     'MAKER':"<![CDATA[{}]]>".format(data['brandNm']),
                     "GOODS_GUBUN":1,
                     "DELV_TYPE":3,
                     "DELV_COST":'3500',
                     "MODEL_NM":"<![CDATA[상세페이지 참조]]>",
                     "BRAND_NM":"<![CDATA[{}]]>".format(data['brandNm'])
                     }}
        }
    else:
        data = {"SABANG_GOODS_REGI": {
            'HEADER': {'SEND_COMPAYNY_ID': 'kaprl010',
                       'SEND_AUTH_KEY': 'AyTNrA19M8M653FX3MJ0CPBY0yG2J4STRbH',
                       'SEND_DATA': timeNow,
                       'SEND_GOODS_CD_RT': 'Y',
                       'RESULT_TYPE': 'XML'},
            'DATA': {'GOODS_NM': "<![CDATA[{}]]>".format(inputElem['상품명']),
                     'GOODS_KEYWORD': "<![CDATA[{}]]>".format(inputElem['상품약어']),
                     'COMPAYNY_GOODS_CD': "<![CDATA[{}]]>".format(str(inputElem['자체상품코드'])),
                     'STATUS': 3,
                     'DPARTNER_ID': "<![CDATA[VISTA{}]]>".format(inputElem['물류처'].split("_")[-1]),
                     'PARTNER_ID': "<![CDATA[vitsonmro]]>",
                     'IMG_PATH': "<![CDATA[{}]]>".format(inputElem['대표이미지']),
                     'GOODS_COST': "<![CDATA[{}]]>".format(inputElem['원가']),
                     'STOCK_USE_YN': 'N',
                     'GOODS_PRICE': "<![CDATA[{}]]>".format(inputElem['판매가']),
                     'GOODS_REMARKS': "<![CDATA[{}]]>".format(inputElem['상품상세설명']),
                     'MAKER': "<![CDATA[{}]]>".format(inputElem['브랜드명']),
                     "GOODS_GUBUN": 1,
                     "DELV_TYPE": 3,
                     "DELV_COST": '3500',
                     "MODEL_NM": "<![CDATA[상세페이지 참조]]>",
                     "BRAND_NM": "<![CDATA[{}]]>".format(inputElem['브랜드명'])
                     }}
        }

    # 변환된 dict를 다시 XML로 변환
    xml_output = dict_to_xml(data)

    # XML 선언과 함께 전체 XML 문자열 생성
    xml_output = '<?xml version="1.0" encoding="euc-kr"?>' + xml_output

    # 파일로 저장
    with open('change.xml', 'w', encoding='euc-kr') as file:
        file.write(xml_output)

    # # XML 파일 불러오기
    # tree = ET.parse('change.xml', parser=ET.XMLParser(encoding='iso-8859-5'))
    # root = tree.getroot()
    #
    #
    #
    #
    # # DATA 태그 내의 SEX 태그 찾기
    # for sex_tag in root.findall('.//HEADER/SEND_DATE'):
    #     # SEX 값 변경
    #     sex_tag.text=timeNow
    #
    # for tag in root.findall('.//DATA/GOODS_NM'):
    #     # SEX 값 변경
    #     tag.text="<![CDATA[{}]]>".format(data['productNm'])
    #
    # for tag in root.findall('.//DATA/GOODS_KEYWORD'):
    #     # SEX 값 변경
    #     tag.text="<![CDATA[{}]]>".format(data['standard'])
    #
    # for tag in root.findall('.//DATA/COMPAYNY_GOODS_CD'):
    #     # SEX 값 변경
    #     tag.text="<![CDATA[vitson_{}]]>".format(str(data['productCode']))
    #
    # for tag in root.findall('.//DATA/STATUS'):
    #     # SEX 값 변경
    #     tag.text=data['balance']
    #
    # for tag in root.findall('.//DATA/DPARTNER_ID'):
    #     # SEX 값 변경
    #     tag.text="<![CDATA[{}]]>".format(data['unit'])
    #
    # for tag in root.findall('.//DATA/IMG_PATH'):
    #     # SEX 값 변경
    #     tag.text="<![CDATA[{}]]>".format(data['pictureNm'])
    #
    # for tag in root.findall('.//DATA/GOODS_COST'):
    #     # SEX 값 변경
    #     tag.text="<![CDATA[{}]]>".format(data['originPrice'])
    #
    # for tag in root.findall('.//DATA/STOCK_USE_YN'):
    #     # SEX 값 변경
    #     tag.text='N'
    #
    # for tag in root.findall('.//DATA/GOODS_PRICE'):
    #     # SEX 값 변경
    #     tag.text="<![CDATA[{}]]>".format(data['salePrice'])
    #
    # for tag in root.findall('.//DATA/GOODS_REMARKS'):
    #     # SEX 값 변경
    #     tag.text="<![CDATA[{}]]>".format(productDetail)
    #
    # # 변경된 XML 파일 저장
    # tree.write('change.xml', encoding='euc-kr')
def dict_to_xml(dict_data):
    """
    사전을 XML 형식의 문자열로 변환하는 함수.
    """
    def build_element(key, value):
        """
        키와 값으로 XML 요소를 구성하는 함수.
        """
        if isinstance(value, dict):
            return f"<{key}>{dict_to_xml(value)}</{key}>"
        else:
            return f"<{key}>{value}</{key}>"

    return ''.join(build_element(key, value) for key, value in dict_data.items())

def UploadImageToS3(file_path):
    # AWS 계정의 액세스 키와 시크릿 키를 설정합니다.
    aws_access_key_id = 'AKIA44DPO5SKIPTFJBTU'
    aws_secret_access_key = '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'

    bucket_name="myvitzonbucket"

    # S3 클라이언트를 생성합니다.
    s3_client = boto3.client('s3', aws_access_key_id=aws_access_key_id, aws_secret_access_key=aws_secret_access_key)
    # print(s3_client)
    # 그림 파일을 S3 버킷에 업로드합니다.
    try:
        s3_client.upload_file(file_path, bucket_name, file_path)
        print("파일 업로드 성공!")
    except Exception as e:
        print("파일 업로드 실패:", e)



while True:
    # # #=============ID가져오기=======
    session=GetLogin()
    with open ('dataList.json', "r",encoding='utf-8-sig') as f:
        dataList = json.load(f)
    dataList.reverse()
    totalList=[]
    timeStart=datetime.datetime.now().strftime("%H%M%S")
    for data in dataList:
        resultList=GetID(data,session)
        totalList.extend(resultList)
        with open('totalList.json', 'w',encoding='utf-8-sig') as f:
            json.dump(totalList, f, indent=2,ensure_ascii=False)
        timeNow=datetime.datetime.now().strftime("%H%M%S")
        print("시작시간:{}/종료시간:{}".format(timeStart,timeNow))


    #===========데이타 불러오기========
    with open ('totalList.json', "r",encoding='utf-8-sig') as f:
        totalList = json.load(f)
    MergeExcel()
    inputList=GetExcel()
    with open('inputList.json', 'w',encoding='utf-8-sig') as f:
        json.dump(inputList, f, indent=2,ensure_ascii=False)

    #================엑셀 기반으로 데이타 반영
    for index,inputElem in enumerate(inputList):
        timeNow=datetime.datetime.now().strftime("%H시%M분%S초")
        productId = inputElem['자체상품코드'].replace("vitson_", "")
        print("{}/{}번째(ID:{}) 확인중..(현재시간:{})".format(index+1,len(inputList),productId,timeNow))
        # print("productId:",productId,"/ productId_TYPE:",type(productId))
        # productId가 product_list의 self_product_code 중 하나와 일치하는지 확인
        # productId와 일치하는 productCode를 가진 요소 찾기
        matchedProduct = None
        for item in totalList:
            if str(item['productCode']) == str(productId):
                matchedProduct = item
                break

        # 결과 출력
        if matchedProduct:
            print("● Matched Product:", matchedProduct)
            statusType=True
            MakeXML(matchedProduct,inputElem,statusType)
            UploadImageToS3('change.xml')
            ChangeProduct()
            print("=====================")
        else:
            statusType = False
            print("X Non-Matched Product:", matchedProduct)
            MakeXML(matchedProduct, inputElem,statusType)
            UploadImageToS3('change.xml')
            ChangeProduct()

# # ================크롤링한 거 기반으로 데이타 반영 모드
# for totalElem in totalList:
#     productId="vitson_"+str(totalElem['productCode'])
#     # print("productId:",productId,"/ productId_TYPE:",type(productId))
#     # productId가 product_list의 self_product_code 중 하나와 일치하는지 확인
#     productDetail = next((product["상품상세설명"] for product in inputList if product["자체상품코드"] == productId), None)
#     # 결과 출력
#     if productDetail:
#         print("일치함")
#         print("totalElem:", totalElem, "/ totalElem_TYPE:", type(totalElem))
#         productNm=totalElem['productNm']
#         # print("일치함 productNm:",productNm,"/ productNm_TYPE:",type(productNm))
#         MakeXML(totalElem,productDetail)
#         UploadImageToS3('change.xml')
#         ChangeProduct()
#         print("=====================")
#     else:
#         print("해당 상품 코드를 가진 상품이 없습니다.")



